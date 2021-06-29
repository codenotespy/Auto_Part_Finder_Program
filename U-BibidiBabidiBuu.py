from tkinter import *
from tkinter import messagebox
from tkinter.scrolledtext import ScrolledText
import csv
import sqlite3
import os
from tkinter.ttk import *
import xlsxwriter
import pandas as pd
import openpyxl


window = Tk()
window.title("U-BibidiBabidiBuu")
window.geometry('1000x550')


def cleartextbox():
    listBox.delete("1.0", "end")


# TO EXPORT EXCEL:
OPTIONS = ["Choose", "Choose", "Export Excel"]  # etc
variable = StringVar(window)
variable.set(OPTIONS[0])  # default value
w = OptionMenu(window, variable, *OPTIONS)
w.place(x=17, y=15)


# To Search by Part No
Partsearch = StringVar()


def searchoe():
    if variable.get() == "Choose":
        listBox.delete("1.0", "end")

        search = Partsearch.get()
        search2 = Partsearch.get()
        search3 = Partsearch.get()
        conn = sqlite3.connect('oe.db')
        cur = conn.cursor()
        cur.execute("SELECT MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], [PRICE(EUR)], [ENGLISH DESCRIPTION], NOTE, MODEL, [MODEL DETAIL], ENGINE, AÇIKLAMA, [PRICE(USD)] FROM OETable WHERE PARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR [OE NO] LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR OLDPARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') COLLATE NOCASE",
            ('%' + search + '%', '%' + search2 + '%', '%' + search3 + '%',))

        rows = cur.fetchall()
        rows.sort(key=lambda e: e[1], reverse=FALSE)
        listBox.insert(END, "POSTION\t|MANUFACTURER\t\t|PART NO \t\t\t|OE NO \t\t|OLD PART NO \t\t\t|PRICE(EUR)\t\t|ENGLISH DESCRIPTION\t\t\t|NOTE\t\t|MODEL\t\t\t|MODEL DETAIL\t\t|ENGINE\t\t|AÇIKLAMA\t\t\t\t|PRICE(USD)")  # "\t\t\t" means like tab btw the collumns
        listBox.insert(END, "\n")

        for i in range(len(rows)):
            listBox.insert(END, (i + 1))  # To show "Position" column
            listBox.insert(END, "\t|")  # "\t\t\t" means like tab btw the collumns
            listBox.insert(END, rows[i][0])  # [0] means show the first column, [i] means show the corresponding rows
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][1])  # [1] means show the second column, [i] means show the corresponding rows
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][2])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][3])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][4])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][5])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][6])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][7])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][8])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][9])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][10])
            listBox.insert(END, "\t\t\t\t|")
            listBox.insert(END, rows[i][11])
            listBox.insert(END, "\n")
        conn.close()

    if variable.get() == "Export Excel":
        search = Partsearch.get()
        search2 = Partsearch.get()
        search3 = Partsearch.get()
        conn = sqlite3.connect('oe.db')
        cur = conn.cursor()
        # TO GET INPUT, SEARCH IN OETable:
        cur.execute("SELECT MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], MODEL, ENGINE, [MODEL DETAIL], [PRICE(USD)], [PRICE(EUR)], NOTE FROM OETable WHERE PARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR [OE NO] LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR OLDPARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') COLLATE NOCASE",
    ('%' + search + '%', '%' + search2 + '%', '%' + search3 + '%',))
        # TO COPY SELECTED DATA IN THE DATABASE
        rows = cur.fetchall()
        # TO CREATE EXCEL FILE
        workbook = xlsxwriter.Workbook('searchoutput.xlsx')
        worksheet = workbook.add_worksheet()
        # TO WRITE IN THE CREATED EXCEL
        worksheet.write('A1', 'MANUFACTURER')
        worksheet.write('B1', 'PART NO')
        worksheet.write('C1', 'OE NO')
        worksheet.write('D1', 'OLD PART NO')
        worksheet.write('E1', 'AÇIKLAMA')
        worksheet.write('F1', 'ENGLISH DESCRIPTION')
        worksheet.write('G1', 'MODEL')
        worksheet.write('H1', 'ENGINE')
        worksheet.write('I1', 'MODEL DETAIL')
        worksheet.write('J1', 'PRICE (USD)')
        worksheet.write('K1', 'PRICE (EUR)')
        worksheet.write('L1', 'NOTE')

        row = 1
        col = 0
        for module in rows:
            worksheet.write_row(row, col, module)
            row += 1

        workbook.close()
        conn.close()
        # TO OPEN THE SAVED EXCEL FILE
        os.system("start EXCEL.EXE searchoutput.xlsx")


def aftermarketsearch():
    if variable.get() == "Choose":
        listBox.delete("1.0", "end")

        search = Partsearch.get()
        search2 = Partsearch.get()
        search3 = Partsearch.get()
        conn = sqlite3.connect('amlist.db')
        cur = conn.cursor()
        cur.execute(
            "SELECT MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], [ENGLISH DESCRIPTION], [PRICE(EUR)], [LIST GROUP], NOTES, MODEL, [MODEL DETAIL], ENGINE, AÇIKLAMA FROM MergedList WHERE PARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR [OENOFORM] LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR OLDPARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') COLLATE NOCASE",
            ('%' + search + '%', '%' + search2 + '%', '%' + search3 + '%',))

        rows = cur.fetchall()
        rows.sort(key=lambda e: e[1], reverse=FALSE)
        listBox.insert(END, "POSTION\t|MANUFACTURER\t\t|PART NO\t\t\t|OE NO \t\t\t|OLD PART NO \t\t\t|ENGLISH DESCRIPTION\t\t\t|PRICE(EUR)\t\tLIST GROUP\t\t|NOTE\t\t|MODEL\t\t\t|MODEL DETAIL\t\t|ENGINE\t\t|AÇIKLAMA")
        # "\t\t\t" means like tab btw the columns
        listBox.insert(END, "\n")
        for i in range(len(rows)):
            listBox.insert(END, (i + 1))  # To show "Position" column
            listBox.insert(END, "\t|")  # "\t\t\t" means like tab btw the columns
            listBox.insert(END, rows[i][0])  # [0] means show the first column, [i] means show the corresponding rows
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][1])  # [1] means show the second column, [i] means show the corresponding rows
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][2])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][3])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][4])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][5])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][6])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][7])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][8])
            listBox.insert(END, "\t\t\t|")
            listBox.insert(END, rows[i][9])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][10])
            listBox.insert(END, "\t\t|")
            listBox.insert(END, rows[i][11])
            listBox.insert(END, "\n")
        conn.close()

    if variable.get() == "Export Excel":
        search = Partsearch.get()
        search2 = Partsearch.get()
        search3 = Partsearch.get()
        conn = sqlite3.connect('amlist.db')
        cur = conn.cursor()
        # TO GET INPUT, SEARCH IN OETable:
        cur.execute("SELECT MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], [ENGLISH DESCRIPTION], AÇIKLAMA, [LIST GROUP], MODEL, [MODEL DETAIL], ENGINE, [PRICE(EUR)], NOTES FROM MergedList WHERE PARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR [OENOFORM] LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') OR OLDPARTNOFORM LIKE REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(?, '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '') COLLATE NOCASE", ('%'+search+'%', '%'+search2+'%', '%'+search3+'%',))
        # TO COPY SELECTED DATA IN THE DATABASE
        rows = cur.fetchall()
        # TO CREATE EXCEL FILE
        workbook = xlsxwriter.Workbook('searchoutput.xlsx')
        worksheet = workbook.add_worksheet()
        # TO WRITE IN THE CREATED EXCEL
        worksheet.write('A1', 'MANUFACTURER')
        worksheet.write('B1', 'PART NO')
        worksheet.write('C1', 'OE NO')
        worksheet.write('D1', 'OLD PART NO')
        worksheet.write('E1', 'ENGLISH DESCRIPTION')
        worksheet.write('F1', 'TURKISH DESCRIPTION')
        worksheet.write('G1', 'LIST GROUP')
        worksheet.write('H1', 'MODEL')
        worksheet.write('I1', 'MODEL DETAIL')
        worksheet.write('J1', 'ENGINE')
        worksheet.write('K1', 'PRICE (EUR)')
        worksheet.write('L1', 'NOTE')

        row = 1
        col = 0
        for module in rows:
            worksheet.write_row(row, col, module)
            row += 1

        workbook.close()
        conn.close()
        # TO OPEN THE SAVED EXCEL FILE
        os.system("start EXCEL.EXE searchoutput.xlsx")


OPTIONS = ["Choose List", "Choose List", "New Cus. Lİst", "Spc. AM. List", "Add AM Stock"]
variable2 = StringVar(window)
variable2.set(OPTIONS[0])  # default value
w = OptionMenu(window, variable2, *OPTIONS)
w.place(x=700, y=15)

OPTIONS2 = ["Choose List", "Choose List", "OE List", "Add OE Stock"]
variable3 = StringVar(window)
variable3.set(OPTIONS2[0])  # default value
w = OptionMenu(window, variable3, *OPTIONS2)
w.place(x=820, y=15)


def updatelists():
    if variable3.get() == "OE List":
        con = sqlite3.connect("oe.db")
        cur = con.cursor()
        # CREATE THE TABLE:
        cur.execute("CREATE TABLE IF NOT EXISTS OETable (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], MODEL, ENGINE, [MODEL DETAIL], [PRICE(USD)], [PRICE(EUR)], NOTE, [QTY IN STOCK]);")  # use your column names here
        # INSERT CSV FILE INTO THE TABLE:
        with open('oeall.csv', 'r', encoding="ISO-8859-1") as fin:  # `with` statement available in 2.5+
            # csv.DictReader uses first line in file for column headings by default
            dr = csv.DictReader(fin)  # comma is default delimiter
            to_db = [(i['MANUFACTURER'], i['PART NO'], i['OE NO'], i['OLD PART NO'], i['AÇIKLAMA'], i['ENGLISH DESCRIPTION'], i['MODEL'], i['ENGINE'], i['MODEL DETAIL'], i['PRICE(USD)'], i['PRICE(EUR)'], i['NOTE'], i['QTY IN STOCK']) for i in dr]
            cur.executemany("INSERT INTO OETable (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], MODEL, ENGINE, [MODEL DETAIL], [PRICE(USD)], [PRICE(EUR)], NOTE, [QTY IN STOCK]) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", to_db)
        con.commit()
        # ADD THE NEW FORM COLUMNS:
        conn = sqlite3.connect('oe.db')
        cur = conn.cursor()
        cur.execute("ALTER TABLE OETable ADD COLUMN PARTNOFORM TEXT")
        cur.execute("ALTER TABLE OETable ADD COLUMN OLDPARTNOFORM TEXT")
        conn.commit()
        # COPY COLUMNS TO THE NEW FORM COLUMNS:
        cur.execute("UPDATE OETable SET PARTNOFORM = [PART NO]")
        cur.execute("UPDATE OETable SET OLDPARTNOFORM = [OLD PART NO]")
        conn.commit()
        # UPDATE FORM COLUMNS FOR SEARCING:
        conn = sqlite3.connect('oe.db')
        cur = conn.cursor()
        cur.execute("UPDATE OETable SET [OLDPARTNOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([OLDPARTNOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")  # Replace all empty characters, commas etc. in OLDPARTNOFORM column
        cur.execute("UPDATE OETable SET [PARTNOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([PARTNOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE OETable SET [PRICE(USD)] = REPLACE([PRICE(USD)], '.', ',')")
        cur.execute("UPDATE OETable SET [PRICE(EUR)] = REPLACE([PRICE(EUR)], '.', ',')")
        conn.commit()
        con.close()
        messagebox.showinfo('You have a message! :)', 'Your OE List has been inserted successfully.')

    if variable3.get() == "Add OE Stock":
        con = sqlite3.connect("oe.db")
        cur = con.cursor()
        cur.execute("DROP TABLE IF EXISTS OEStockList")
        con.commit()
        # CREATE THE TABLE:
        cur.execute("CREATE TABLE IF NOT EXISTS OEStockList (MANUFACTURER, [PART NO], [QTY IN STOCK]);")  # use your column names here
        # INSERT CSV FILE INTO THE TABLE:
        with open('oestocklist.csv', 'r', encoding="ISO-8859-1") as fin:  # `with` statement available in 2.5+
            # csv.DictReader uses first line in file for column headings by default
            dr = csv.DictReader(fin)  # comma is default delimiter
            to_db = [(i['MANUFACTURER'], i['PART NO'], i['QTY IN STOCK']) for i in dr]
            cur.executemany("INSERT INTO OEStockList (MANUFACTURER, [PART NO], [QTY IN STOCK]) VALUES (?, ?, ?);", to_db)
        con.commit()

        ### TO MAKE VLOOKUP AND MATCHES STOCKS FROM OEStockList Table WITH OETable and INSERT THE RESULTS IN "OETABLEWITHSTOCK" Table
        cur.execute("DROP TABLE IF EXISTS OETABLEWITHSTOCK")
        con.commit()
        cur.execute("CREATE TABLE IF NOT EXISTS OETABLEWITHSTOCK (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], [AÇIKLAMA], [ENGLISH DESCRIPTION], MODEL, ENGINE, [MODEL DETAIL], [PRICE(USD)], [PRICE(EUR)], NOTE, [QTY IN STOCK], PARTNOFORM, OLDPARTNOFORM)")
        con.commit()
        cur.execute("INSERT INTO OETABLEWITHSTOCK SELECT OETable.MANUFACTURER, OETable.[PART NO], OETable.[OE NO], OETable.[OLD PART NO], OETable.[AÇIKLAMA], OETable.[ENGLISH DESCRIPTION], OETable.MODEL, OETable.ENGINE, OETable.[MODEL DETAIL], OETable.[PRICE(USD)], OETable.[PRICE(EUR)], OETable.NOTE, OEStockList.[QTY IN STOCK], OETable.PARTNOFORM, OETable.OLDPARTNOFORM FROM OETable LEFT JOIN OEStockList ON OETable.[PART NO]=OEStockList.[PART NO]")
        con.commit()
        con.close()
        # TO RENAME PREVIOUS OETable with "Previous-OETable" And REPLACE "OETABLEWITHSTOCK" TABLE NAME WITH "OETable"
        con2 = sqlite3.connect("oe.db")
        cur2 = con2.cursor()
        cur2.execute("DROP TABLE IF EXISTS PreviousOETable")
        con2.commit()
        cur2.execute("ALTER TABLE OETable RENAME TO PreviousOETable")
        con2.commit()
        cur2.execute("ALTER TABLE OETABLEWITHSTOCK RENAME TO OETable")
        con2.commit()
        # cur2.execute("SELECT IFNULL([QTY IN STOCK], '0') FROM OETable")
        con2.close()
        messagebox.showinfo('You have a message! :)', 'Your OE Stock List has been inserted successfully.')
    if variable2.get() == "New Cus. Lİst":
        con = sqlite3.connect("amlist.db")
        cur = con.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS NewCustomerAMList (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], [LIST GROUP], MODEL, ENGINE, [MODEL DETAIL], [PRICE(EUR)], NOTES, [QTY IN STOCK]);")
        with open('newcustam.csv', 'r', encoding="ISO-8859-1") as fin:
            dr2 = csv.DictReader(fin)
            to_db2 = [(i['MANUFACTURER'], i['PART NO'], i['OE NO'], i['OLD PART NO'], i['AÇIKLAMA'], i['ENGLISH DESCRIPTION'], i['LIST GROUP'], i['MODEL'], i['ENGINE'], i['MODEL DETAIL'], i['PRICE(EUR)'], i['NOTES'], i['QTY IN STOCK']) for i in dr2]
            cur.executemany("INSERT INTO NewCustomerAMList (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], [LIST GROUP], MODEL, ENGINE, [MODEL DETAIL], [PRICE(EUR)], NOTES, [QTY IN STOCK]) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", to_db2)
            con.commit()
        conn = sqlite3.connect('amlist.db')
        cur = conn.cursor()
        cur.execute("ALTER TABLE NewCustomerAMList ADD COLUMN PARTNOFORM TEXT")
        cur.execute("ALTER TABLE NewCustomerAMList ADD COLUMN OLDPARTNOFORM TEXT")
        cur.execute("ALTER TABLE NewCustomerAMList ADD COLUMN OENOFORM TEXT")
        conn.commit()
        cur.execute("UPDATE NewCustomerAMList SET PARTNOFORM = [PART NO]")
        cur.execute("UPDATE NewCustomerAMList SET OLDPARTNOFORM = [OLD PART NO]")
        cur.execute("UPDATE NewCustomerAMList SET OENOFORM = [OE NO]")
        conn.commit()
        conn = sqlite3.connect('amlist.db')
        cur = conn.cursor()
        cur.execute("UPDATE NewCustomerAMList SET [OLDPARTNOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([OLDPARTNOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE NewCustomerAMList SET [PARTNOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([PARTNOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE NewCustomerAMList SET [OENOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([OENOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE NewCustomerAMList SET [PRICE(EUR)] = REPLACE([PRICE(EUR)], '.', ',')")
        conn.commit()
        con.close()
        messagebox.showinfo('You have a message! :)', 'New Customer Aftermarket List has been inserted successfully.')

    if variable2.get() == "Spc. AM. List":
        con = sqlite3.connect("amlist.db")
        cur = con.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS SpecialAMList (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], [LIST GROUP], MODEL, ENGINE, [MODEL DETAIL], [PRICE(EUR)], NOTES, [QTY IN STOCK]);")
        with open('specialamlist.csv', 'r', encoding="ISO-8859-1") as fin:
            dr2 = csv.DictReader(fin)
            to_db2 = [(i['MANUFACTURER'], i['PART NO'], i['OE NO'], i['OLD PART NO'], i['AÇIKLAMA'], i['ENGLISH DESCRIPTION'], i['LIST GROUP'], i['MODEL'], i['ENGINE'], i['MODEL DETAIL'], i['PRICE(EUR)'], i['NOTES'], i['QTY IN STOCK']) for i in dr2]
            cur.executemany("INSERT INTO SpecialAMList (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], [LIST GROUP], MODEL, ENGINE, [MODEL DETAIL], [PRICE(EUR)], NOTES, [QTY IN STOCK]) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", to_db2)
            con.commit()
        conn = sqlite3.connect('amlist.db')
        cur = conn.cursor()
        cur.execute("ALTER TABLE SpecialAMList ADD COLUMN PARTNOFORM TEXT")
        cur.execute("ALTER TABLE SpecialAMList ADD COLUMN OLDPARTNOFORM TEXT")
        cur.execute("ALTER TABLE SpecialAMList ADD COLUMN OENOFORM TEXT")
        cur.execute("UPDATE SpecialAMList SET OENOFORM = [OE NO]")
        cur.execute("UPDATE SpecialAMList SET PARTNOFORM = [PART NO]")
        cur.execute("UPDATE SpecialAMList SET OLDPARTNOFORM = [OLD PART NO]")
        conn.commit()
        conn = sqlite3.connect('amlist.db')
        cur = conn.cursor()
        cur.execute("UPDATE SpecialAMList SET [OLDPARTNOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([OLDPARTNOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE SpecialAMList SET [PARTNOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([PARTNOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE SpecialAMList SET [OENOFORM] = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([OENOFORM], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")
        cur.execute("UPDATE SpecialAMList SET [PRICE(EUR)] = REPLACE([PRICE(EUR)], '.', ',')")
        conn.commit()
        con.close()
        messagebox.showinfo('You have a message! :)', 'Special Aftermarket List has been inserted successfully.')

    if variable2.get() == "Add AM Stock":
        con = sqlite3.connect("amlist.db")
        cur = con.cursor()
        cur.execute("DROP TABLE IF EXISTS AMStockList")
        con.commit()
        # CREATE THE TABLE:
        cur.execute("CREATE TABLE IF NOT EXISTS AMStockList (MANUFACTURER, [PART NO], [QTY IN STOCK]);")  # use your column names here
        # INSERT CSV FILE INTO THE TABLE:
        with open('amstocklist.csv', 'r', encoding="ISO-8859-1") as fin:  # `with` statement available in 2.5+
        # csv.DictReader uses first line in file for column headings by default
            dr = csv.DictReader(fin)  # comma is default delimiter
            to_db = [(i['MANUFACTURER'], i['PART NO'], i['QTY IN STOCK']) for i in dr]
            cur.executemany("INSERT INTO AMStockList (MANUFACTURER, [PART NO], [QTY IN STOCK]) VALUES (?, ?, ?);", to_db)
            con.commit()

        ### TO MAKE VLOOKUP AND MATCHES STOCKS FROM OEStockList Table WITH OETable and INSERT THE RESULTS IN "OETABLEWITHSTOCK" Table
        cur.execute("DROP TABLE IF EXISTS AMTABLEWITHSTOCK")
        con.commit()
        cur.execute("CREATE TABLE IF NOT EXISTS AMTABLEWITHSTOCK (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], [AÇIKLAMA], [ENGLISH DESCRIPTION], [LIST GROUP], MODEL, ENGINE, [MODEL DETAIL], [PRICE(EUR)], NOTES, [QTY IN STOCK], PARTNOFORM, OLDPARTNOFORM, OENOFORM)")
        con.commit()
        # LEFT JOIN BRINGS ALL FROM LEFT TABLE('MergedList') and BRINGS ALL [QTY IN STOCK] FROM RIGHT TABLE('AMStockList') FOR THE ROWS THAT PART NO COLLUMNS MATCHES FOR BOTH TABLES.
        cur.execute("INSERT INTO AMTABLEWITHSTOCK SELECT MergedList.MANUFACTURER, MergedList.[PART NO], MergedList.[OE NO], MergedList.[OLD PART NO], MergedList.[AÇIKLAMA], MergedList.[ENGLISH DESCRIPTION], MergedList.[LIST GROUP], MergedList.MODEL, MergedList.ENGINE, MergedList.[MODEL DETAIL], MergedList.[PRICE(EUR)], MergedList.NOTES, AMStockList.[QTY IN STOCK], MergedList.PARTNOFORM, MergedList.OLDPARTNOFORM, MergedList.OENOFORM FROM MergedList LEFT JOIN AMStockList ON MergedList.[PART NO]=AMStockList.[PART NO]")
        con.commit()
        con.close()
        # TO RENAME PREVIOUS OETable with "Previous-OETable" And REPLACE "OETABLEWITHSTOCK" TABLE NAME WITH "OETable"
        con2 = sqlite3.connect("amlist.db")
        cur2 = con2.cursor()
        cur2.execute("DROP TABLE IF EXISTS PreviousMergedList")
        con2.commit()
        cur2.execute("ALTER TABLE MergedList RENAME TO PreviousMergedList")
        con2.commit()
        cur2.execute("ALTER TABLE AMTABLEWITHSTOCK RENAME TO MergedList")
        con2.commit()
        con2.close()
        messagebox.showinfo('You have a message! :)', 'Your AM Stock List has been inserted successfully.')


def deletedb():
    if variable3.get() == "OE List":
        sqlite3.connect("oe.db")
        os.remove("oe.db")
        messagebox.showinfo('You have a message! :)', 'Your OE list has been deleted successfully.')

    if variable2.get() == "New Cus. Lİst":
        sqlite3.connect("amlist.db")
        os.remove("amlist.db")
        messagebox.showinfo('You have a message! :)', 'Your New Customer Aftermarket List has been deleted successfully.')

    if variable2.get() == "Spc. AM. List":
        sqlite3.connect("amlist.db")
        os.remove("amlist.db")
        messagebox.showinfo('You have a message! :)', 'Your Special Aftermarket List has been deleted successfully.')


def mergeamlists():
    con = sqlite3.connect("amlist.db")
    cur = con.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS MergedList (MANUFACTURER, [PART NO], [OE NO], [OLD PART NO], AÇIKLAMA, [ENGLISH DESCRIPTION], [LIST GROUP], MODEL, ENGINE, [MODEL DETAIL], [PRICE(EUR)], NOTES, [QTY IN STOCK], PARTNOFORM, OLDPARTNOFORM, OENOFORM);")
    con.commit()
    cur.execute("INSERT INTO MergedList SELECT * FROM SpecialAMList;")
    cur.execute("INSERT INTO MergedList SELECT * FROM NewCustomerAMList;")
    con.commit()
    con.close()
    messagebox.showinfo('You have a message! :)', 'Your List is ready to use.')


def openexc():
    os.system("start EXCEL.EXE multiplesearchx.xlsx")


def expmultipleoeprice():
    con = sqlite3.connect("oe.db")
    cur = con.cursor()

    # To delete table itself:
    cur.execute("DROP TABLE IF EXISTS OETableMultiple")
    con.commit()

    # To create the table
    cur.execute("CREATE TABLE IF NOT EXISTS OETableMultiple ([PART NO], PARTNOFORM)")
    con.commit()

    # To insert data from excel into the table:
    con = sqlite3.connect('oe.db')
    wb = pd.ExcelFile('multiplesearchx.xlsx')
    for sheet in wb.sheet_names:
        df = pd.read_excel('multiplesearchx.xlsx', sheet_name='Sheet1')
        df.to_sql('OETableMultiple', con, index=False, if_exists="append")
    con.commit()

    # UPDATE FORM COLUMNS FOR SEARCING:
    con = sqlite3.connect('oe.db')
    cur = con.cursor()
    cur.execute("UPDATE OETableMultiple SET PARTNOFORM = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([PART NO], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")  # Replace all empty characters, commas etc. in OLDPARTNOFORM column
    con.commit()

    # To join:
    con = sqlite3.connect("oe.db")
    cur = con.cursor()
    cur.execute("SELECT OETableMultiple.[PART NO], OETable.MANUFACTURER, OETable.[PART NO], OETable.[OE NO], OETable.[OLD PART NO], OETable.[AÇIKLAMA], OETable.[ENGLISH DESCRIPTION], OETable.MODEL, OETable.ENGINE, OETable.[MODEL DETAIL], OETable.[PRICE(USD)], OETable.[PRICE(EUR)], OETable.NOTE, OETable.[QTY IN STOCK] FROM OETableMultiple LEFT JOIN OETable ON OETable.[OE NO]=OETableMultiple.PARTNOFORM OR OETable.PARTNOFORM=OETableMultiple.PARTNOFORM OR OETable.OLDPARTNOFORM=OETableMultiple.PARTNOFORM")

    rows = cur.fetchall()
    # TO CREATE EXCEL FILE
    workbook = xlsxwriter.Workbook('searchoutput.xlsx')
    worksheet = workbook.add_worksheet()
    # TO WRITE IN THE CREATED EXCEL
    worksheet.write('A1', 'ENQ')
    worksheet.write('B1', 'MANUFACTURER')
    worksheet.write('C1', 'PART NO')
    worksheet.write('D1', 'OE NO')
    worksheet.write('E1', 'OLD PART NO')
    worksheet.write('F1', 'AÇIKLAMA')
    worksheet.write('G1', 'ENGLISH DESCRIPTION')
    worksheet.write('H1', 'MODEL')
    worksheet.write('I1', 'ENGINE')
    worksheet.write('J1', 'MODEL DETAIL')
    worksheet.write('K1', 'PRICE (USD)')
    worksheet.write('L1', 'PRICE (EUR)')
    worksheet.write('M1', 'NOTE')
    worksheet.write('N1', 'QTY IN STOCK')

    row = 1
    col = 0
    for module in rows:
        worksheet.write_row(row, col, module)
        row += 1

    workbook.close()
    con.close()
    # TO OPEN THE SAVED EXCEL FILE
    os.system("start EXCEL.EXE searchoutput.xlsx")


def expmultipleamprice():
    con = sqlite3.connect("amlist.db")
    cur = con.cursor()

    # To delete table itself:
    cur.execute("DROP TABLE IF EXISTS AMTableMultiple")
    con.commit()

    # To create the table
    cur.execute("CREATE TABLE IF NOT EXISTS AMTableMultiple ([PART NO], PARTNOFORM)")
    con.commit()

    # To insert data from excel into the table:
    con = sqlite3.connect('amlist.db')
    wb = pd.ExcelFile('multiplesearchx.xlsx')
    for sheet in wb.sheet_names:
        df = pd.read_excel('multiplesearchx.xlsx', sheet_name='Sheet1')
        df.to_sql('AMTableMultiple', con, index=False, if_exists="append")
    con.commit()

        # UPDATE FORM COLUMNS FOR SEARCING:
    con = sqlite3.connect('amlist.db')
    cur = con.cursor()
    cur.execute("UPDATE AMTableMultiple SET PARTNOFORM = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([PART NO], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")  # Replace all empty characters, commas etc. in OLDPARTNOFORM column
    con.commit()

    # To join:
    con = sqlite3.connect("amlist.db")
    cur = con.cursor()
    cur.execute("SELECT AMTableMultiple.[PART NO], MergedList.MANUFACTURER, MergedList.[PART NO], MergedList.[OE NO], MergedList.[OLD PART NO], MergedList.AÇIKLAMA, MergedList.[ENGLISH DESCRIPTION], MergedList.[LIST GROUP], MergedList.MODEL, MergedList.ENGINE, MergedList.[MODEL DETAIL], MergedList.NOTES, MergedList.[PRICE(EUR)], MergedList.[QTY IN STOCK] FROM AMTableMultiple LEFT JOIN MergedList ON MergedList.PARTNOFORM LIKE '%' ||AMTableMultiple.PARTNOFORM|| '%'")

    rows = cur.fetchall()
    # TO CREATE EXCEL FILE
    workbook = xlsxwriter.Workbook('searchoutput.xlsx')
    worksheet = workbook.add_worksheet()
    # TO WRITE IN THE CREATED EXCEL
    worksheet.write('A1', 'ENQ. NO')
    worksheet.write('B1', 'MANUFACTURER')
    worksheet.write('C1', 'PART NO')
    worksheet.write('D1', 'OE NO')
    worksheet.write('E1', 'OLD PART NO')
    worksheet.write('F1', 'AÇIKLAMA')
    worksheet.write('G1', 'ENGLISH DESCRIPTION')
    worksheet.write('H1', 'LIST GROUP')
    worksheet.write('I1', 'MODEL')
    worksheet.write('J1', 'ENGINE')
    worksheet.write('K1', 'MODEL DETAIL')
    worksheet.write('L1', 'NOTES')
    worksheet.write('M1', 'PRICE (EUR)')
    worksheet.write('N1', 'QTY IN STOCK')


    row = 1
    col = 0
    for module in rows:
        worksheet.write_row(row, col, module)
        row += 1

    workbook.close()
    con.close()
    # TO OPEN THE SAVED EXCEL FILE
    os.system("start EXCEL.EXE searchoutput.xlsx")


def expmultipleamversion():
    con = sqlite3.connect("amlist.db")
    cur = con.cursor()

    # To delete table itself:
    cur.execute("DROP TABLE IF EXISTS AMTableMultiple")
    con.commit()

    # To create the table
    cur.execute("CREATE TABLE IF NOT EXISTS AMTableMultiple ([PART NO], PARTNOFORM)")
    con.commit()

    # To insert data from excel into the table:
    con = sqlite3.connect('amlist.db')
    wb = pd.ExcelFile('multiplesearchx.xlsx')
    for sheet in wb.sheet_names:
        df = pd.read_excel('multiplesearchx.xlsx', sheet_name='Sheet1')
        df.to_sql('AMTableMultiple', con, index=False, if_exists="append")
    con.commit()


        # UPDATE FORM COLUMNS FOR SEARCING:
    con = sqlite3.connect('amlist.db')
    cur = con.cursor()
    cur.execute("UPDATE AMTableMultiple SET PARTNOFORM = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE([PART NO], '-', ''), ',', ''), ' ', ''), '.', ''), '/', ''), '\\', ''), '|', ''), '*', ''), '+', '')")  # Replace all empty characters, commas etc. in OLDPARTNOFORM column
    con.commit()


    # To join:
    con = sqlite3.connect("amlist.db")
    cur = con.cursor()
    cur.execute("SELECT AMTableMultiple.[PART NO], MergedList.MANUFACTURER, MergedList.[PART NO], MergedList.[PRICE(EUR)], MergedList.[QTY IN STOCK], MergedList.[OE NO], MergedList.NOTES FROM AMTableMultiple LEFT JOIN MergedList ON MergedList.OENOFORM LIKE '%' ||AMTableMultiple.PARTNOFORM|| '%'")
    # MergedList.OENOFORM LIKE '%' ||AMTableMultiple.PARTNOFORM|| '%'

    rows = cur.fetchall()
    # TO CREATE EXCEL FILE
    workbook = xlsxwriter.Workbook('searchoutput.xlsx')
    worksheet = workbook.add_worksheet()
    # TO WRITE IN THE CREATED EXCEL
    worksheet.write('A1', 'ENQ. NO')
    worksheet.write('B1', 'MANUFACTURER')
    worksheet.write('C1', 'PART NO')
    worksheet.write('D1', 'PRICE (EUR)')
    worksheet.write('E1', 'QTY IN STOCK')
    worksheet.write('F1', 'OE NO')
    worksheet.write('G1', 'NOTES')

    row = 1
    col = 0
    for module in rows:
        worksheet.write_row(row, col, module)
        row += 1
    workbook.close()
    con.close()
    # We've finished transfering data from database to excel named 'searchoutput.xlsx'. Now we will copy data from this excel to 'U-searchoutput.xlsx' below: ---

        ### PART 2= TO COPY & PASTE FROM THE FILE TO U-searchoutput.xlsx & From multiplesearchx.xlsx
    # Before Copy & Paste, We should clear values from previous copy paste..
    wb2 = openpyxl.load_workbook("U-searchoutput.xlsx")
    Sheet1 = wb2['list']
    Sheet2 = wb2['U-List']
    # To clear Specific Range of Values in sheet named 'list'
    for ax in Sheet1['B2:F10000']:
        for cell in ax:
            cell.value = None
    # To clear values in sheet named 'U-List'
    for bx in Sheet2['A2:A10000']:
        for cell in bx:
            cell.value = None
    # TO COMMIT & SAVE THE CLEARED FILE
    wb2.save("U-searchoutput.xlsx")


        ### TO CONTINUE COPY & PASTE FROM THE FILE TO U-searchoutput.xlsx
    # File to be copied from:
    wb = openpyxl.load_workbook("searchoutput.xlsx")
    Sheet = wb['Sheet1']

    # File-2 to be copied from:
    wb3 = openpyxl.load_workbook("multiplesearchx.xlsx")
    Sheet3 = wb3['Sheet1']

    # File to be pasted into (We don't have to write below because we already wrote this above to clear the file):
    wb2 = openpyxl.load_workbook("U-searchoutput.xlsx")
    Sheet1 = wb2['list']
    Sheet2 = wb2['U-List']

    # TO COPY FROM "searchoutput.xlsx" FILE TO 'list' sheet in U-searchoutput.xlsx
    for r in range(1, 10000):
        for c in range(1, 8):
            Sheet1.cell(row=r,column=c+1).value = Sheet.cell(row=r,column=c).value
            # I wrote 'c+1' to start from column B instead of column A

    # TO COPY FROM "multiplesearchx.xlsx" FILE TO 'U-List' sheet in U-searchoutput.xlsx
    for r in range(1, 10000):
        for c in range(1, 2):
            Sheet2.cell(row=r+1,column=c).value = Sheet3.cell(row=r+1,column=c).value

    # TO COMMIT & SAVE THE FILE
    wb2.save("U-searchoutput.xlsx")
    # TO OPEN THE SAVED EXCEL FILE
    os.system("start EXCEL.EXE U-searchoutput.xlsx")


btn = Button(window, text="Clear", width=11, command=cleartextbox)
btn.place(x=190, y=15)

btn = Button(window, text="Search OE", width=11, command=searchoe)
btn.place(x=190, y=45)

btn = Button(window, text="Replacem.", width=11, command=aftermarketsearch)
btn.place(x=270, y=45)

btn = Button(window, text="Insert Exc..", width=11, command=openexc)
btn.place(x=400, y=15)

btn = Button(window, text="OE Price", width=11, command=expmultipleoeprice)
btn.place(x=400, y=45)

btn = Button(window, text="AM Price", width=11, command=expmultipleamprice)
btn.place(x=480, y=45)

btn = Button(window, text="AM Versions", width=11, command=expmultipleamversion)
btn.place(x=480, y=15)

btn = Button(window, text="Update", width=9, command=updatelists)
btn.place(x=630, y=15)

btn = Button(window, text="Delete", width=9, command=deletedb)
btn.place(x=630, y=45)

btn = Button(window, text="Merge the Lists", width=15, command=mergeamlists)
btn.place(x=700, y=45)


# Horizontal (x) Scroll bar
xscrollbar = Scrollbar(window, orient=HORIZONTAL)
xscrollbar.pack(side=BOTTOM, fill=X)
# Vertical (y) Scroll Bar
yscrollbar = Scrollbar(window)
yscrollbar.pack(side=RIGHT, fill=Y)
# Text Widget
listBox = Text(window, wrap=NONE, width=160, xscrollcommand=xscrollbar.set, yscrollcommand=yscrollbar.set)
listBox.place(x=5, y=120)
# Configure the scrollbars
xscrollbar.config(command=listBox.xview)
yscrollbar.config(command=listBox.yview)


entry_1 = Entry(window, textvar=Partsearch)
entry_1.place(x=15, y=45, width=170, height=25)

window.mainloop()